import openpyxl

wb = openpyxl.load_workbook("Base de Datos.xlsx", data_only=True)
print("Sheet Names:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\nAudit Sheet: {name}")
    print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
    
    # Check if there is any data by printing the first few non-empty rows/cells found
    non_empty_rows = []
    for r in range(1, min(ws.max_row + 1, 150)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(v is not None for v in row_vals):
            non_empty_rows.append((r, row_vals))
            
    print(f"Found {len(non_empty_rows)} non-empty rows (scanned up to row 150)")
    if non_empty_rows:
        print("Sample non-empty rows:")
        for r_num, vals in non_empty_rows[:10]:
            # Print index and first few non-null elements
            clean_vals = [str(v)[:30] if v is not None else "" for v in vals[:8]]
            print(f"Row {r_num}: {clean_vals}")
